
from django.forms import modelform_factory
from django.http import HttpResponse
from openpyxl import Workbook
from django.shortcuts import render, get_object_or_404, redirect


# Create your views here.
from estudiante.form import PersonaFormulario
from estudiante.models import Estudiante


def detalle_estudiante(request,id):
    # persona = Persona.objects.get(pk= id)
    estudiante = get_object_or_404(Estudiante, pk=id)
    mensaje = {'estudiante': estudiante}
    return render(request, 'detalle.html', mensaje)

#PersonaFormulario = modelform_factory(Estudiante, exclude=[])
def nuevo_estudiante(request):
    if request.POST:
        formEstudiante = PersonaFormulario(request.POST)
        if formEstudiante.is_valid():
            formEstudiante.save()
            return redirect('inicio')
    else:
        formEstudiante = PersonaFormulario()
    return render(request, 'nuevo.html', {'formEstudiante': formEstudiante})

def modificar_estudiante(request, id):
    estudiante = get_object_or_404(Estudiante, pk=id)
    if request.method == 'POST':
        formEstudiante = PersonaFormulario(request.POST, instance= estudiante)
        if formEstudiante.is_valid():
            formEstudiante.save()
            return redirect('inicio')
    else:
        formEstudiante = PersonaFormulario(instance=estudiante)
    mensaje = {'formEstudiante': formEstudiante}
    return render(request, 'modificar.html', context = mensaje)

def eliminar_estudiante(request, id):
    estudiante = get_object_or_404(Estudiante, pk=id)
    if estudiante:
        estudiante.delete()
    return redirect('inicio')

def reporte_estudiantes(request):
    # Obtenemos todas las personas de nuestra base de datos
    estudiantes = Estudiante.objects.order_by('apellido')
    # Creamos el libro de trabajo
    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
    ws['B1'] = 'REPORTE DE ESTUDIANTES'
    # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
    ws.merge_cells('B1:O1')
    ws.title = 'Reporte'
    # Creamos los encabezados desde la celda B3 hasta la E3
    ws['B3'] = 'ID'
    ws['C3'] = 'NOMBRE'
    ws['D3'] = 'APELLIDO'
    ws['E3'] = 'GENERO'
    ws['F3'] = 'EMAIL ESTUDIANTE'
    ws['G3'] = 'CIUDAD'
    ws['H3'] = 'CURSO'
    ws['I3'] = 'AULA'
    ws['J3'] = 'SEMESTRE'
    ws['K3'] = 'MODALIDAD'
    ws['L3'] = 'NIVEL'
    ws['M3'] = 'TUTOR ASIGNADO'
    ws['N3'] = 'TITULO'
    ws['O3'] = 'EMAIL TUTOR'
    cont = 4
    # Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
    for estudiante in estudiantes:
        ws.cell(row=cont, column=2).value = estudiante.id
        ws.cell(row=cont, column=3).value = estudiante.nombre
        ws.cell(row=cont, column=4).value = estudiante.apellido
        ws.cell(row=cont, column=5).value = estudiante.genero
        ws.cell(row=cont, column=6).value = estudiante.email
        ws.cell(row=cont, column=7).value = estudiante.direccion.calle
        ws.cell(row=cont, column=8).value = estudiante.curso.nombrecurso
        ws.cell(row=cont, column=9).value = estudiante.curso.aula
        ws.cell(row=cont, column=10).value = estudiante.curso.semestre
        ws.cell(row=cont, column=11).value = estudiante.curso.modalidad
        ws.cell(row=cont, column=12).value = estudiante.curso.nivel
        ws.cell(row=cont, column=13).value = estudiante.tutor.nombretutor
        ws.cell(row=cont, column=14).value = estudiante.tutor.titulo
        ws.cell(row=cont, column=15).value = estudiante.tutor.email
        cont = cont + 1
    # Establecemos el nombre del archivo
    nombre_archivo = "ReporteEstudiantes.xlsx"
    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response